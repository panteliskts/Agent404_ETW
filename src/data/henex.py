from __future__ import annotations

import io
import re
import zipfile
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl
import pandas as pd
import requests

from config import GR_TIMEZONE, RAW_DIR

BASE = "https://www.enexgroup.gr"
LIVE_URL = f"{BASE}/web/guest/markets-publications-el-day-ahead-market"
ARCHIVE_URL = f"{BASE}/web/guest/dam-idm-archive"
HEADERS = {"User-Agent": "Mozilla/5.0 (battery-opt-hackathon)"}

HENEX_DIR = RAW_DIR / "henex"
HENEX_DIR.mkdir(parents=True, exist_ok=True)
ZIPS_DIR = HENEX_DIR / "zips"
ZIPS_DIR.mkdir(parents=True, exist_ok=True)
XLSX_DIR = HENEX_DIR / "xlsx"
XLSX_DIR.mkdir(parents=True, exist_ok=True)

DOC_TYPES_DEFAULT = ("ResultsSummary", "Results", "AggrCurves", "PreMarketSummary")

SECTION_HEADERS = {
    "Total SELL Trades": "volume_sell",
    "Total BUY Trades": "volume_buy",
    "Market Clearing Price": "price",
    "PRODUCTION TECHNOLOGY / MTU": "production",
    "DEMAND / MTU": "demand",
}

PRICE_LABELS = {
    "Greece Mainland  (15min MCP)": "dam_price_eur_mwh",
    "Greece Mainland (60min Index)": "dam_price_60min_idx_eur_mwh",
    "Greece Mainland": "dam_price_eur_mwh",
}

VOLUME_LABELS = {
    "Greece Mainland": "volume_mainland_mwh",
}

GEN_LABELS = {
    "LIGNITE": "gen_lignite_mw",
    "GAS": "gen_gas_mw",
    "HYDRO": "gen_hydro_mw",
    "RENEWABLES": "gen_renewables_mw",
    "CRETE RENEWABLES": "gen_crete_renewables_mw",
    "CRETE CONVENTIONAL": "gen_crete_conventional_mw",
    "BESS": "gen_bess_mw",
    "PRODUCTION": "production_total_mw",
}

LOAD_LABELS = {
    "HV LOAD": "load_hv_mw",
    "MV LOAD": "load_mv_mw",
    "LV LOAD": "load_lv_mw",
    "PUMP": "load_pump_mw",
    "SYSTEM LOSSES": "system_losses_mw",
    "CRETE LOAD": "load_crete_mw",
    "D/R LOAD": "load_demand_response_mw",
    "DEMAND": "demand_total_mw",
    "BESS": "load_bess_mw",
    "RENEWABLES": "renewables_buy_mw",
}

_LIVE_LINK_RE = re.compile(
    r'href="(/c/document_library/get_file\?uuid=[0-9a-f-]+&groupId=20126)"[^>]*>'
    r"\s*(?:<i[^>]*></i>)?\s*&nbsp;\s*"
    r"([^<]+\.xlsx)",
    re.S,
)
_ARCHIVE_LINK_RE = re.compile(
    r'href="(/c/document_library/get_file\?uuid=[0-9a-f-]+&groupId=20126)"[^>]*>'
    r"\s*(?:<i[^>]*></i>)?\s*&nbsp;\s*"
    r"(\d{4}_EL[^<]+_(?:Results|ResultsSummary|AggrCurves|PreMarketSummary|PrelimResults|POSNOMs|NDPS|BLKORDRs|MWO)\.zip)",
    re.S,
)


def _get(url: str) -> requests.Response:
    r = requests.get(url, headers=HEADERS, timeout=60)
    r.raise_for_status()
    return r


def list_recent_files() -> list[tuple[str, str]]:
    text = _get(LIVE_URL).text
    return [(BASE + url, fn.strip()) for url, fn in _LIVE_LINK_RE.findall(text)]


def list_archive_zips() -> list[tuple[str, str]]:
    text = _get(ARCHIVE_URL).text
    return [(BASE + url, fn.strip()) for url, fn in _ARCHIVE_LINK_RE.findall(text)]


def _doc_type_of(filename: str) -> str | None:
    parts = filename.split("_")
    for token in parts:
        for t in (
            "ResultsSummary",
            "PrelimResults",
            "Results",
            "AggrCurves",
            "PreMarketSummary",
            "POSNOMs",
            "NDPS",
            "BLKORDRs",
            "MWO",
        ):
            if token.startswith(t):
                return t
    return None


def download_archives(years: list[int], doc_types: tuple[str, ...] = DOC_TYPES_DEFAULT) -> list[Path]:
    zips = list_archive_zips()
    out = []
    for url, fn in zips:
        m = re.match(r"(\d{4})_", fn)
        if not m:
            continue
        year = int(m.group(1))
        if year not in years:
            continue
        dt = _doc_type_of(fn)
        if dt not in doc_types:
            continue
        path = ZIPS_DIR / fn
        if not path.exists():
            print(f"  downloading {fn} ...")
            r = _get(url)
            path.write_bytes(r.content)
        else:
            print(f"  cached {fn}")
        out.append(path)
    return out


def extract_zip(zip_path: Path, dest: Path = XLSX_DIR, dam_only: bool = True) -> list[Path]:
    extracted = []
    with zipfile.ZipFile(zip_path) as zf:
        for name in zf.namelist():
            if not name.lower().endswith(".xlsx"):
                continue
            base = Path(name).name
            if dam_only and "EL-DAM_" not in base:
                continue
            target = dest / base
            if not target.exists():
                with zf.open(name) as src, open(target, "wb") as dst:
                    dst.write(src.read())
            extracted.append(target)
    return extracted


def _latest_version_per_date(paths: list[Path]) -> list[Path]:
    by_key: dict[str, Path] = {}
    for p in paths:
        m = re.match(r"(\d{8})_EL-DAM_(\w+?)_EN_v(\d+)\.xlsx", p.name)
        if not m:
            continue
        key = f"{m.group(1)}_{m.group(2)}"
        version = int(m.group(3))
        if key not in by_key or version > int(re.search(r"v(\d+)", by_key[key].name).group(1)):
            by_key[key] = p
    return list(by_key.values())


def download_recent(doc_types: tuple[str, ...] = DOC_TYPES_DEFAULT) -> list[Path]:
    out = []
    for url, fn in list_recent_files():
        dt = _doc_type_of(fn)
        if dt not in doc_types:
            continue
        path = XLSX_DIR / fn
        if not path.exists():
            print(f"  downloading {fn} ...")
            r = _get(url)
            path.write_bytes(r.content)
        out.append(path)
    return out


def _date_from_filename(fn: str) -> datetime | None:
    m = re.match(r"(\d{8})_", fn)
    if not m:
        return None
    return datetime.strptime(m.group(1), "%Y%m%d")


def _find_sheets(wb: openpyxl.Workbook) -> tuple[str | None, str | None]:
    sell = buy = None
    for sn in wb.sheetnames:
        u = sn.upper()
        if "SELL" in u and sell is None:
            sell = sn
        elif "BUY" in u and buy is None:
            buy = sn
    return sell, buy


def parse_results_summary(xlsx_path: Path) -> pd.DataFrame:
    """Parse one ResultsSummary xlsx into a tidy DataFrame indexed by 15-min (or hourly) timestamps in Europe/Athens."""
    date = _date_from_filename(xlsx_path.name)
    if date is None:
        raise ValueError(f"cannot parse date from {xlsx_path.name}")

    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    sell_sheet, buy_sheet = _find_sheets(wb)
    if sell_sheet is None and buy_sheet is None:
        raise ValueError(f"no summary sheets in {xlsx_path.name}: {wb.sheetnames}")

    mtu_count = _detect_mtu_count(wb[sell_sheet or buy_sheet])
    delta_min = 15 if mtu_count == 96 else 60
    start = datetime(date.year, date.month, date.day)
    timestamps = pd.DatetimeIndex(
        [start + timedelta(minutes=delta_min * i) for i in range(mtu_count)]
    )

    records: dict[str, list] = {}

    if sell_sheet:
        _harvest_sheet(
            wb[sell_sheet], mtu_count, records,
            section_label_maps={
                "volume_sell": VOLUME_LABELS,
                "price": PRICE_LABELS,
                "production": GEN_LABELS,
            },
        )
    if buy_sheet:
        _harvest_sheet(
            wb[buy_sheet], mtu_count, records,
            section_label_maps={
                "price": PRICE_LABELS,
                "demand": LOAD_LABELS,
            },
            allow_overwrite=False,
        )

    if "dam_price_eur_mwh" not in records:
        raise ValueError(f"no DAM price row found in {xlsx_path.name}")

    df = pd.DataFrame(records, index=timestamps)
    df = df.apply(pd.to_numeric, errors="coerce")
    df.index = _localize_athens(df.index)
    df.index.name = "time"
    return df


def _detect_mtu_count(ws) -> int:
    best = 0
    for r in range(1, 6):
        row = next(ws.iter_rows(min_row=r, max_row=r, values_only=True), None)
        if not row:
            continue
        ints = []
        for v in row[1:]:
            if isinstance(v, (int, float)):
                ints.append(int(v))
            elif isinstance(v, str) and v.strip().isdigit():
                ints.append(int(v.strip()))
        if not ints:
            continue
        mx = max(ints)
        if mx in (24, 96) and mx > best and ints[:5] == [1, 2, 3, 4, 5]:
            best = mx
    return best or 96


def _harvest_sheet(ws, mtu_count, records, section_label_maps, allow_overwrite=False):
    current_section = None
    for row in ws.iter_rows(values_only=True):
        if not row or row[0] is None:
            continue
        label = str(row[0]).strip()
        if label in SECTION_HEADERS:
            current_section = SECTION_HEADERS[label]
            continue
        label_map = section_label_maps.get(current_section)
        if label_map is None or label not in label_map:
            continue
        col_name = label_map[label]
        if col_name in records and not allow_overwrite:
            continue
        values = list(row[1 : 1 + mtu_count])
        if all(v is None for v in values):
            continue
        records[col_name] = values


def _localize_athens(idx: pd.DatetimeIndex) -> pd.DatetimeIndex:
    """Localize naive Athens-local timestamps, gracefully handling DST gaps and duplicates."""
    try:
        return idx.tz_localize(GR_TIMEZONE, ambiguous="infer", nonexistent="shift_forward")
    except Exception:
        flags = [True] * len(idx)
        seen = set()
        for i, ts in enumerate(idx):
            key = (ts.month, ts.day, ts.hour, ts.minute)
            if key in seen:
                flags[i] = False
            else:
                seen.add(key)
        return idx.tz_localize(GR_TIMEZONE, ambiguous=flags, nonexistent="shift_forward")


def build_dataset(years: list[int], include_recent: bool = True) -> pd.DataFrame:
    print(f"[henex] downloading archives for years {years} ...")
    zips = download_archives(years, doc_types=("ResultsSummary",))
    xlsx_files: list[Path] = []
    for z in zips:
        xlsx_files.extend(extract_zip(z))
    if include_recent:
        print("[henex] fetching recent files from live page ...")
        xlsx_files.extend(download_recent(doc_types=("ResultsSummary",)))

    xlsx_files = [p for p in xlsx_files if "ResultsSummary" in p.name]
    xlsx_files = _latest_version_per_date(xlsx_files)
    print(f"[henex] parsing {len(xlsx_files)} xlsx files ...")
    frames = []
    for p in sorted(xlsx_files):
        try:
            frames.append(parse_results_summary(p))
        except Exception as exc:
            print(f"  skip {p.name}: {exc}")
    if not frames:
        return pd.DataFrame()
    df = pd.concat(frames).sort_index()
    df = df[~df.index.duplicated(keep="last")]
    return df


def save(years: list[int]) -> Path:
    df = build_dataset(years)
    if df.empty:
        raise RuntimeError("HEnEx returned no data")
    path = RAW_DIR / f"henex_results_{years[0]}_{years[-1]}.parquet"
    df.to_parquet(path)
    print(f"  saved {path.name}  rows={len(df)}  cols={len(df.columns)}")
    return path
